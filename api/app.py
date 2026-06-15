"""
Módulo: api/app.py
API REST do Nexus construída com Flask.
Expõe os recursos: /alunos, /turmas, /indicadores, /relatorios
"""

import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from flask import Flask, request, jsonify, render_template, redirect, url_for, send_file
import io
import hashlib
import secrets
import uuid
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from services.database import init_db, get_connection, close_connection
from services.repositorio_sqlite import RepositorioSQLite
from services.aluno_service import AlunoService
from services.indicador_service import IndicadorService
from services.lyceum_service import LyceumService
from models.usuario import Usuario, PerfilUsuario
from models.aluno import Aluno
from services.bokeh_service import BokehService
from api.rbac import anonimizar_dados
from services.database import get_connection
import re
import urllib.parse

# ------------------------------------------------------------------ #
# Inicialização
# ------------------------------------------------------------------ #


# Tokens de sessão ativos {token: usuario_id} - Mantido globalmente para testes
_tokens_ativos: dict = {}

def create_app(config=None):
    """Cria e configura uma instância da aplicação Flask."""
    app = Flask(__name__)
    if config:
        app.config.update(config)
    @app.teardown_appcontext
    def teardown_db(exception):
        close_connection(exception)

    with app.app_context():
        init_db()

    # Repositórios
    repo_alunos      = RepositorioSQLite("alunos")
    repo_turmas      = RepositorioSQLite("disciplinas")
    repo_indicadores = RepositorioSQLite("indicadores")
    repo_usuarios    = RepositorioSQLite("usuarios")
    repo_auditoria   = RepositorioSQLite("auditoria")

    # Serviços
    aluno_service     = AlunoService(repo_alunos)
    indicador_service = IndicadorService(repo_indicadores)
    lyceum_service    = LyceumService()

    # Tokens de sessão ativos {token: usuario_id}
    # Utiliza o dicionário global _tokens_ativos

    # ------------------------------------------------------------------ #
    # Autenticação simples (token no header)
    # ------------------------------------------------------------------ #

    def requer_autenticacao(f):
        """Decorator: exige token válido no header Authorization."""
        @wraps(f)
        def decorado(*args, **kwargs):
            token = request.headers.get("Authorization", "").replace("Bearer ", "")
            if token not in _tokens_ativos:
                return jsonify({"erro": "Não autorizado. Token inválido ou ausente."}), 401
            return f(*args, **kwargs)
        return decorado


    def requer_permissao(acao: str):
        """Decorator: verifica se o usuário tem permissão para a ação."""
        def decorator(f):
            @wraps(f)
            def decorado(*args, **kwargs):
                token = request.headers.get("Authorization", "").replace("Bearer ", "")
                usuario_id = _tokens_ativos.get(token)
                dados_usuario = repo_usuarios.buscar(usuario_id)
                if not dados_usuario:
                    return jsonify({"erro": "Usuário não encontrado."}), 403
                perfil = PerfilUsuario(dados_usuario["perfil"])
                usuario = Usuario(
                    id=dados_usuario["id"],
                    nome=dados_usuario["nome"],
                    matricula=dados_usuario.get("matricula", ""),
                    perfil=perfil,
                )
                if not usuario.tem_permissao(acao):
                    return jsonify({"erro": f"Permissão negada para ação: {acao}"}), 403
                return f(*args, **kwargs)
            return decorado
        return decorator


    # ------------------------------------------------------------------ #
    # Rotas de interface gráfica (Frontend)
    # ------------------------------------------------------------------ #

    @app.route("/login", methods=["GET"])
    def login_page():
        """
        GET /login
        Retorna a página HTML de login.
        """
        return render_template("login.html")


    @app.route("/dashboard", methods=["GET"])
    def dashboard_page():
        """
        GET /dashboard
        Retorna a página de Dashboard analítico.
        """
        return render_template("dashboard.html")


    @app.route("/cadastro", methods=["GET"])
    def cadastro_page():
        """
        GET /cadastro
        Retorna a página de cadastro de novo usuário.
        """
        return render_template("cadastro.html")


    # ------------------------------------------------------------------ #
    # Rotas de autenticação
    # ------------------------------------------------------------------ #

    @app.route("/auth/login", methods=["POST"])
    def login():
        """
        POST /auth/login
        Body: { "matricula": "...", "senha": "..." }
        Retorna token de acesso.
        """
        dados = request.get_json()
        if not dados or "matricula" not in dados or "senha" not in dados:
            return jsonify({"erro": "Matrícula e senha são obrigatórios."}), 400

        resultado = repo_usuarios.filtrar("matricula", dados["matricula"])
        if not resultado:
            return jsonify({"erro": "Credenciais inválidas."}), 401

        dados_usuario = resultado[0]
        senha_hash = hashlib.sha256(dados["senha"].encode()).hexdigest()
        if dados_usuario.get("senha_hash") != senha_hash:
            return jsonify({"erro": "Credenciais inválidas."}), 401

        token = secrets.token_hex(32)
        _tokens_ativos[token] = dados_usuario["id"]
        return jsonify({"token": token, "perfil": dados_usuario["perfil"], "nome": dados_usuario["nome"]}), 200


    @app.route("/auth/registrar", methods=["POST"])
    def registrar():
        """
        POST /auth/registrar
        Body: { "nome", "matricula", "email", "senha", "cursos": [...] }
        Cria um novo usuário professor. Rota pública.
        """
        dados = request.get_json() or {}
        for campo in ["nome", "matricula", "email", "senha"]:
            if not dados.get(campo):
                return jsonify({"erro": f"Campo obrigatório ausente: {campo}"}), 400

        cursos = dados.get("cursos", [])
        if not cursos:
            return jsonify({"erro": "Selecione ao menos um curso."}), 400

        if repo_usuarios.filtrar("matricula", dados["matricula"]):
            return jsonify({"erro": "Matrícula já cadastrada."}), 409

        novo_id = str(uuid.uuid4())
        repo_usuarios.salvar(novo_id, {
            "id": novo_id,
            "nome": dados["nome"],
            "matricula": dados["matricula"],
            "email": dados["email"],
            "perfil": "professor",
            "cursos": cursos,
            "senha_hash": hashlib.sha256(dados["senha"].encode()).hexdigest(),
            "ativo": True,
        })

        return jsonify({"mensagem": "Conta criada com sucesso."}), 201


    @app.route("/auth/logout", methods=["POST"])
    @requer_autenticacao
    def logout():
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        _tokens_ativos.pop(token, None)
        return jsonify({"mensagem": "Logout realizado com sucesso."}), 200


    # ------------------------------------------------------------------ #
    # Rotas de alunos  /alunos
    # ------------------------------------------------------------------ #

    @app.route("/alunos", methods=["GET"])
    @requer_autenticacao
    def listar_alunos():
        """GET /alunos — lista todos os alunos, com filtros opcionais por periodo, risco e curso."""
        periodo      = request.args.get("periodo", type=int)
        risco_filtro = request.args.get("risco")
        curso_filtro = request.args.get("curso")

        if periodo:
            alunos = aluno_service.listar_por_periodo(periodo)
        else:
            alunos = aluno_service.listar_todos()

        if curso_filtro:
            alunos = [a for a in alunos if a.curso.lower() == curso_filtro.lower()]

        if risco_filtro:
            periodo_risco = periodo or 1
            alunos = [
                a for a in alunos
                if indicador_service.classificar_risco(a, periodo_risco).nivel.value == risco_filtro
            ]

        return jsonify([a.to_dict() for a in alunos]), 200


    @app.route("/alunos/<aluno_id>", methods=["GET"])
    @requer_autenticacao
    def buscar_aluno(aluno_id):
        """GET /alunos/<id> — retorna um aluno específico."""
        try:
            aluno = aluno_service.buscar_por_id(aluno_id)
            return jsonify(aluno.to_dict()), 200
        except ValueError as e:
            return jsonify({"erro": str(e)}), 404


    @app.route("/alunos", methods=["POST"])
    @requer_autenticacao
    @requer_permissao("escrita")
    def cadastrar_aluno():
        """
        POST /alunos
        Body: { "nome", "matricula", "curso", "periodo", "email"(opcional) }
        """
        dados = request.get_json()
        campos_obrigatorios = ["nome", "matricula", "curso", "periodo"]
        for campo in campos_obrigatorios:
            if campo not in dados:
                return jsonify({"erro": f"Campo obrigatório ausente: {campo}"}), 400
        try:
            aluno = aluno_service.cadastrar(**dados)
            token = request.headers.get("Authorization", "").replace("Bearer ", "")
            registrar_auditoria("cadastrar_aluno", token, {"matricula": dados["matricula"]})
            return jsonify(aluno.to_dict()), 201
        except ValueError as e:
            return jsonify({"erro": str(e)}), 409


    @app.route("/alunos/<aluno_id>/notas", methods=["POST"])
    @requer_autenticacao
    @requer_permissao("escrita")
    def registrar_nota(aluno_id):
        """
        POST /alunos/<id>/notas
        Body: { "disciplina", "nota", "periodo" }
        """
        dados = request.get_json()
        try:
            aluno = aluno_service.registrar_nota(
                aluno_id,
                dados["disciplina"],
                float(dados["nota"]),
                int(dados["periodo"]),
            )
            token = request.headers.get("Authorization", "").replace("Bearer ", "")
            registrar_auditoria("registrar_nota", token, {"aluno_id": aluno_id, "disciplina": dados["disciplina"]})
            return jsonify(aluno.to_dict()), 200
        except (ValueError, KeyError) as e:
            return jsonify({"erro": str(e)}), 400


    @app.route("/alunos/<aluno_id>/frequencias", methods=["POST"])
    @requer_autenticacao
    @requer_permissao("escrita")
    def registrar_frequencia(aluno_id):
        """
        POST /alunos/<id>/frequencias
        Body: { "disciplina", "percentual", "periodo" }
        """
        dados = request.get_json()
        try:
            aluno = aluno_service.registrar_frequencia(
                aluno_id,
                dados["disciplina"],
                float(dados["percentual"]),
                int(dados["periodo"]),
            )
            token = request.headers.get("Authorization", "").replace("Bearer ", "")
            registrar_auditoria("registrar_frequencia", token, {"aluno_id": aluno_id})
            return jsonify(aluno.to_dict()), 200
        except (ValueError, KeyError) as e:
            return jsonify({"erro": str(e)}), 400


    # ------------------------------------------------------------------ #
    # Integração Lyceum
    # ------------------------------------------------------------------ #

    @app.route("/sync/lyceum", methods=["POST"])
    @requer_autenticacao
    def sincronizar_lyceum():
        """
        POST /sync/lyceum
        Body: { "url", "usuario", "senha" }
        """
        return jsonify({"aviso": "Integração com Lyceum ainda não implementada.", "status": "pendente"}), 501

        dados = request.get_json()
        if not all(k in dados for k in ("url", "usuario", "senha")):
            return jsonify({"erro": "URL, usuário e senha são obrigatórios."}), 400

        sucesso, mensagem = lyceum_service.autenticar(dados["url"], dados["usuario"], dados["senha"])
        if not sucesso:
            return jsonify({"erro": mensagem}), 401

        # Extrai os dados (Simulado por enquanto, até termos o parse real)
        novas_notas = lyceum_service.extrair_notas(dados["url"])

        # Exemplo de como isso atualizaria o banco (lógica a ser refinada)
        # for nota in novas_notas:
        #     aluno_service.registrar_nota(...)

        return jsonify({"mensagem": "Sincronização concluída com sucesso (Modo: Estrutura)", "sucesso": True}), 200


    # ------------------------------------------------------------------ #
    # Relatórios
    # ------------------------------------------------------------------ #

    @app.route("/relatorios/turma", methods=["GET"])
    @requer_autenticacao
    @requer_permissao("relatorios")
    def relatorio_turma():
        """GET /relatorios/turma?periodo=1 — relatório consolidado por aluno."""
        periodo = request.args.get("periodo", 1, type=int)
        alunos = aluno_service.listar_por_periodo(periodo)
        todos = alunos
        relatorio = []
        for aluno in alunos:
            media = indicador_service.calcular_media_geral(aluno, periodo)
            freq  = indicador_service.calcular_frequencia_media(aluno, periodo)
            risco     = indicador_service.classificar_risco(aluno, periodo)
            zscore    = indicador_service.calcular_zscore_aluno(todos, aluno.id, periodo)
            percentil = indicador_service.calcular_percentil_aluno(todos, aluno.id, periodo)
            iaa       = indicador_service.calcular_iaa(aluno, periodo)
            irp       = indicador_service.calcular_irp(aluno, periodo)
            relatorio.append({
                "aluno_id":        aluno.id,
                "nome":            aluno.nome,
                "matricula":       aluno.matricula,
                "media_geral":     media,
                "frequencia_media": freq,
                "iaa":             iaa,
                "irp":             irp,
                "risco":           risco.nivel.value,
                "zscore":          zscore,
                "percentil":       percentil,
            })
        return jsonify({"periodo": periodo, "total": len(relatorio), "alunos": relatorio}), 200


    @app.route("/relatorios/estatisticas", methods=["GET"])
    @requer_autenticacao
    @requer_permissao("relatorios")
    def estatisticas_turma():
        """GET /relatorios/estatisticas?periodo=1 — estatísticas descritivas da turma."""
        periodo = request.args.get("periodo", 1, type=int)
        alunos = aluno_service.listar_por_periodo(periodo)
        stats = indicador_service.calcular_estatisticas_turma(alunos, periodo)
        correlacao = indicador_service.calcular_correlacao_freq_nota(alunos, periodo)
        distribuicao = indicador_service.calcular_distribuicao_medias(alunos, periodo)
        pontos = indicador_service.calcular_pontos_dispersao(alunos, periodo)
        return jsonify({
            "periodo": periodo,
            "estatisticas_turma": stats,
            "correlacao_freq_nota": correlacao,
            "interpretacao_correlacao": _interpretar_correlacao(correlacao),
            "distribuicao_medias": distribuicao,
            "pontos_dispersao": pontos,
        }), 200


    @app.route("/relatorios/disciplinas", methods=["GET"])
    @requer_autenticacao
    @requer_permissao("relatorios")
    def estatisticas_disciplinas():
        """GET /relatorios/disciplinas?periodo=1 — estatísticas por disciplina."""
        periodo = request.args.get("periodo", 1, type=int)
        alunos = aluno_service.listar_por_periodo(periodo)
        stats = indicador_service.calcular_estatisticas_por_disciplina(alunos, periodo)
        return jsonify({"periodo": periodo, "disciplinas": stats}), 200


    # --- Métricas aceitas no ranking ---
    _METRICAS_RANKING = {"iaa", "irp", "media_geral", "frequencia_media", "percentil", "zscore"}


    @app.route("/relatorios/ranking", methods=["GET"])
    @requer_autenticacao
    @requer_permissao("relatorios")
    def ranking_turma():
        """GET /relatorios/ranking?periodo=1&metrica=iaa&ordem=desc — ranking de alunos por métrica."""
        periodo = request.args.get("periodo", 1, type=int)
        metrica = request.args.get("metrica", "iaa")
        ordem   = request.args.get("ordem", "desc")

        if metrica not in _METRICAS_RANKING:
            return jsonify({"erro": f"Métrica inválida. Use: {', '.join(sorted(_METRICAS_RANKING))}"}), 400
        if ordem not in ("asc", "desc"):
            return jsonify({"erro": "Parâmetro 'ordem' deve ser 'asc' ou 'desc'."}), 400

        alunos = aluno_service.listar_por_periodo(periodo)
        todos  = alunos

        entradas = []
        for aluno in alunos:
            if metrica == "iaa":
                valor = indicador_service.calcular_iaa(aluno, periodo)
            elif metrica == "irp":
                valor = indicador_service.calcular_irp(aluno, periodo)
            elif metrica == "media_geral":
                valor = indicador_service.calcular_media_geral(aluno, periodo)
            elif metrica == "frequencia_media":
                valor = indicador_service.calcular_frequencia_media(aluno, periodo)
            elif metrica == "percentil":
                valor = indicador_service.calcular_percentil_aluno(todos, aluno.id, periodo)
            else:  # zscore
                valor = indicador_service.calcular_zscore_aluno(todos, aluno.id, periodo)
            entradas.append({"aluno_id": aluno.id, "nome": aluno.nome,
                              "matricula": aluno.matricula, "valor": valor})

        entradas.sort(key=lambda e: e["valor"], reverse=(ordem == "desc"))
        for i, e in enumerate(entradas, 1):
            e["posicao"] = i

        return jsonify({
            "periodo": periodo, "metrica": metrica, "ordem": ordem,
            "total": len(entradas), "ranking": entradas,
        }), 200


    @app.route("/alertas/tendencia", methods=["GET"])
    @requer_autenticacao
    def alertas_tendencia():
        """GET /alertas/tendencia?periodo_atual=2 — detecta quedas de IAA entre períodos consecutivos."""
        periodo_atual    = request.args.get("periodo_atual", 2, type=int)
        periodo_anterior = periodo_atual - 1

        alunos  = aluno_service.listar_todos()
        alertas = []

        for aluno in alunos:
            periodos_notas = {n["periodo"] for n in aluno.get_notas()}
            if periodo_atual not in periodos_notas or periodo_anterior not in periodos_notas:
                continue

            iaa_ant = indicador_service.calcular_iaa(aluno, periodo_anterior)
            iaa_atu = indicador_service.calcular_iaa(aluno, periodo_atual)
            variacao = round(iaa_atu - iaa_ant, 2)

            if variacao <= -1.0:
                nivel = "alto"
            elif variacao <= -0.5:
                nivel = "medio"
            else:
                continue

            alertas.append({
                "nivel":                nivel,
                "aluno_id":             aluno.id,
                "aluno_nome":           aluno.nome,
                "tipo":                 "tendencia_queda",
                "iaa_periodo_anterior": iaa_ant,
                "iaa_periodo_atual":    iaa_atu,
                "variacao":             variacao,
                "mensagem":             f"Queda de {abs(variacao)} pontos no IAA entre os períodos {periodo_anterior} e {periodo_atual}",
                "acao":                 "Verificar causas da queda e acionar suporte pedagógico",
            })

        return jsonify({
            "periodo_atual":    periodo_atual,
            "periodo_anterior": periodo_anterior,
            "total":            len(alertas),
            "alertas":          alertas,
        }), 200


    @app.route("/relatorios/cursos", methods=["GET"])
    @requer_autenticacao
    @requer_permissao("relatorios")
    def relatorio_cursos():
        """GET /relatorios/cursos?periodo=1 — resumo consolidado de desempenho por curso."""
        periodo = request.args.get("periodo", 1, type=int)
        alunos  = aluno_service.listar_por_periodo(periodo)

        # Agrupa alunos e notas por curso
        cursos: dict = {}
        for aluno in alunos:
            c = aluno.curso
            cursos.setdefault(c, {"alunos": [], "disciplinas_notas": {}})
            cursos[c]["alunos"].append(aluno)
            for nota in aluno.get_notas():
                if nota["periodo"] == periodo:
                    disc = nota["disciplina"]
                    cursos[c]["disciplinas_notas"].setdefault(disc, [])
                    cursos[c]["disciplinas_notas"][disc].append(nota["nota"])

        resultado = []
        for curso, dados in sorted(cursos.items()):
            lista_alunos = dados["alunos"]
            medias = [indicador_service.calcular_media_geral(a, periodo) for a in lista_alunos]
            freqs  = [indicador_service.calcular_frequencia_media(a, periodo) for a in lista_alunos]
            riscos = [indicador_service.classificar_risco(a, periodo).nivel.value for a in lista_alunos]

            medias_val = [m for m in medias if m > 0]
            freqs_val  = [f for f in freqs  if f > 0]

            disc_medias = {
                disc: round(sum(ns) / len(ns), 2)
                for disc, ns in dados["disciplinas_notas"].items() if ns
            }
            pior_disciplina = min(disc_medias, key=disc_medias.get) if disc_medias else None

            resultado.append({
                "curso":                  curso,
                "total_alunos":           len(lista_alunos),
                "media_turma":            round(sum(medias_val) / len(medias_val), 2) if medias_val else 0.0,
                "frequencia_media_turma": round(sum(freqs_val)  / len(freqs_val),  2) if freqs_val  else 0.0,
                "alunos_risco_alto":      riscos.count("alto"),
                "alunos_risco_medio":     riscos.count("medio"),
                "alunos_risco_baixo":     riscos.count("baixo"),
                "pior_disciplina":        pior_disciplina,
            })

        return jsonify({"periodo": periodo, "cursos": resultado}), 200


    @app.route("/relatorios/turma/exportar", methods=["GET"])
    @requer_autenticacao
    @requer_permissao("relatorios")
    def exportar_relatorio_turma_excel():
        """GET /relatorios/turma/exportar?periodo=1 — exporta o relatório de turma em Excel."""
        periodo = request.args.get("periodo", 1, type=int)
        alunos = aluno_service.listar_por_periodo(periodo)
        todos = alunos

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Turma - Período {periodo}"

        cabecalho = ["Nome", "Matrícula", "Média Geral", "Frequência (%)", "IAA", "IRP", "Z-Score", "Percentil", "Risco"]
        ws.append(cabecalho)

        # Estilo do cabeçalho
        azul = PatternFill("solid", fgColor="2563EB")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = azul
            cell.alignment = Alignment(horizontal="center")

        cores_risco = {"alto": "FCA5A5", "medio": "FDE68A", "baixo": "BBF7D0"}

        for aluno in alunos:
            media     = indicador_service.calcular_media_geral(aluno, periodo)
            freq      = indicador_service.calcular_frequencia_media(aluno, periodo)
            iaa       = indicador_service.calcular_iaa(aluno, periodo)
            irp       = indicador_service.calcular_irp(aluno, periodo)
            zscore    = indicador_service.calcular_zscore_aluno(todos, aluno.id, periodo)
            percentil = indicador_service.calcular_percentil_aluno(todos, aluno.id, periodo)
            risco     = indicador_service.classificar_risco(aluno, periodo)

            linha = [
                aluno.nome,
                aluno.matricula,
                round(media, 2) if media is not None else "",
                round(freq, 1) if freq is not None else "",
                round(iaa, 3) if iaa is not None else "",
                round(irp, 3) if irp is not None else "",
                round(zscore, 3) if zscore is not None else "",
                round(percentil, 1) if percentil is not None else "",
                risco.nivel.value,
            ]
            ws.append(linha)

            # Cor na célula de risco (última coluna)
            nivel = risco.nivel.value
            if nivel in cores_risco:
                ws.cell(row=ws.max_row, column=9).fill = PatternFill("solid", fgColor=cores_risco[nivel])

        # Ajuste de largura das colunas
        larguras = [30, 14, 13, 16, 10, 10, 10, 10, 10]
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"relatorio_turma_periodo{periodo}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


    @app.route("/relatorios/alunos/exportar", methods=["GET"])
    @requer_autenticacao
    @requer_permissao("relatorios")
    def exportar_alunos_excel():
        """GET /relatorios/alunos/exportar?periodo=1 — exporta lista de alunos com notas e frequências em Excel."""
        periodo = request.args.get("periodo", type=int)
        alunos = aluno_service.listar_por_periodo(periodo) if periodo else aluno_service.listar_todos()

        wb = openpyxl.Workbook()

        # Aba: Alunos
        ws_alunos = wb.active
        ws_alunos.title = "Alunos"
        cab_alunos = ["Nome", "Matrícula", "Curso", "Período", "Email"]
        ws_alunos.append(cab_alunos)
        azul = PatternFill("solid", fgColor="2563EB")
        for cell in ws_alunos[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = azul
            cell.alignment = Alignment(horizontal="center")
        for aluno in alunos:
            ws_alunos.append([aluno.nome, aluno.matricula, aluno.curso, aluno.periodo, aluno.email or ""])
        for i, w in enumerate([30, 14, 25, 10, 30], 1):
            ws_alunos.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Aba: Notas
        ws_notas = wb.create_sheet("Notas")
        cab_notas = ["Nome", "Matrícula", "Disciplina", "Nota", "Período"]
        ws_notas.append(cab_notas)
        for cell in ws_notas[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = azul
            cell.alignment = Alignment(horizontal="center")
        for aluno in alunos:
            for n in aluno.get_notas():
                ws_notas.append([aluno.nome, aluno.matricula, n["disciplina"], n["nota"], n["periodo"]])
        for i, w in enumerate([30, 14, 25, 8, 10], 1):
            ws_notas.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Aba: Frequências
        ws_freq = wb.create_sheet("Frequências")
        cab_freq = ["Nome", "Matrícula", "Disciplina", "Frequência (%)", "Período"]
        ws_freq.append(cab_freq)
        for cell in ws_freq[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = azul
            cell.alignment = Alignment(horizontal="center")
        for aluno in alunos:
            for f in aluno.get_frequencias():
                ws_freq.append([aluno.nome, aluno.matricula, f["disciplina"], f["percentual"], f["periodo"]])
        for i, w in enumerate([30, 14, 25, 14, 10], 1):
            ws_freq.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        nome_arquivo = f"alunos_periodo{periodo}.xlsx" if periodo else "alunos_todos.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


    def registrar_auditoria(acao: str, token: str, detalhes: dict = {}):
        """Persiste um registro de auditoria com usuário, ação e timestamp."""
        from datetime import datetime
        usuario_id  = _tokens_ativos.get(token, "desconhecido")
        entrada_id  = str(uuid.uuid4())
        repo_auditoria.salvar(entrada_id, {
            "id":         entrada_id,
            "usuario_id": usuario_id,
            "acao":       acao,
            "detalhe":    detalhes,
            "timestamp":  datetime.now().isoformat(),
        })


    def _interpretar_correlacao(r: float) -> str:
        """Retorna texto interpretativo para o coeficiente de correlação de Pearson."""
        abs_r = abs(r)
        direcao = "positiva" if r >= 0 else "negativa"
        if abs_r >= 0.9:
            forca = "muito forte"
        elif abs_r >= 0.7:
            forca = "forte"
        elif abs_r >= 0.5:
            forca = "moderada"
        elif abs_r >= 0.3:
            forca = "fraca"
        else:
            forca = "muito fraca ou inexistente"
        return f"Correlação {direcao} {forca} (r={r})"


    # ------------------------------------------------------------------ #
    # Painel Admin — Gestão de Usuários
    # ------------------------------------------------------------------ #

    @app.route("/admin/usuarios", methods=["GET"])
    @requer_autenticacao
    @requer_permissao("usuarios")
    def admin_listar_usuarios():
        usuarios = repo_usuarios.listar()
        return jsonify([{
            "id": u["id"],
            "nome": u["nome"],
            "matricula": u["matricula"],
            "perfil": u["perfil"],
            "ativo": u.get("ativo", True),
        } for u in usuarios]), 200


    @app.route("/admin/usuarios", methods=["POST"])
    @requer_autenticacao
    @requer_permissao("usuarios")
    def admin_criar_usuario():
        dados = request.get_json() or {}
        campos = ["nome", "matricula", "senha", "perfil"]
        for c in campos:
            if c not in dados:
                return jsonify({"erro": f"Campo obrigatório ausente: {c}"}), 400
        perfis_validos = ["admin", "professor", "coordenador", "visualizador"]
        if dados["perfil"] not in perfis_validos:
            return jsonify({"erro": f"Perfil inválido. Use: {', '.join(perfis_validos)}"}), 400
        existente = repo_usuarios.filtrar("matricula", dados["matricula"])
        if existente:
            return jsonify({"erro": "Matrícula já cadastrada."}), 409
        novo_id = str(uuid.uuid4())
        usuario = {
            "id": novo_id,
            "nome": dados["nome"],
            "matricula": dados["matricula"],
            "perfil": dados["perfil"],
            "senha_hash": hashlib.sha256(dados["senha"].encode()).hexdigest(),
            "ativo": True,
        }
        repo_usuarios.salvar(novo_id, usuario)
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        registrar_auditoria("criar_usuario", token, {"matricula": dados["matricula"], "perfil": dados["perfil"]})
        return jsonify({"id": novo_id, "nome": usuario["nome"], "matricula": usuario["matricula"], "perfil": usuario["perfil"], "ativo": True}), 201


    @app.route("/admin/usuarios/<usuario_id>", methods=["PUT"])
    @requer_autenticacao
    @requer_permissao("usuarios")
    def admin_atualizar_usuario(usuario_id):
        dados = request.get_json() or {}
        usuario = repo_usuarios.buscar(usuario_id)
        if not usuario:
            return jsonify({"erro": "Usuário não encontrado."}), 404
        if "perfil" in dados:
            perfis_validos = ["admin", "professor", "coordenador", "visualizador"]
            if dados["perfil"] not in perfis_validos:
                return jsonify({"erro": "Perfil inválido."}), 400
            usuario["perfil"] = dados["perfil"]
        if "ativo" in dados:
            usuario["ativo"] = bool(dados["ativo"])
        if "senha" in dados and dados["senha"]:
            usuario["senha_hash"] = hashlib.sha256(dados["senha"].encode()).hexdigest()
        if "nome" in dados:
            usuario["nome"] = dados["nome"]
        repo_usuarios.salvar(usuario_id, usuario)
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        registrar_auditoria("atualizar_usuario", token, {"usuario_id": usuario_id})
        return jsonify({"mensagem": "Usuário atualizado.", "id": usuario_id}), 200


    @app.route("/admin/usuarios/<usuario_id>", methods=["DELETE"])
    @requer_autenticacao
    @requer_permissao("usuarios")
    def admin_deletar_usuario(usuario_id):
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        if _tokens_ativos.get(token) == usuario_id:
            return jsonify({"erro": "Você não pode remover sua própria conta."}), 400
        removido = repo_usuarios.deletar(usuario_id)
        if not removido:
            return jsonify({"erro": "Usuário não encontrado."}), 404
        registrar_auditoria("remover_usuario", token, {"usuario_id": usuario_id})
        return jsonify({"mensagem": "Usuário removido."}), 200


    @app.route("/admin/auditoria", methods=["GET"])
    @requer_autenticacao
    @requer_permissao("usuarios")
    def admin_auditoria():
        """GET /admin/auditoria?limit=50 — retorna as últimas entradas do log de auditoria."""
        limit   = request.args.get("limit", 50, type=int)
        entradas = repo_auditoria.listar()
        entradas.sort(key=lambda e: e.get("timestamp", ""), reverse=True)
        return jsonify(entradas[:limit]), 200


    # ------------------------------------------------------------------ #
    # Alertas Inteligentes
    # ------------------------------------------------------------------ #

    @app.route("/alertas", methods=["GET"])
    @requer_autenticacao
    def gerar_alertas():
        """GET /alertas?periodo=2 — detecta padrões de risco e gera alertas automáticos."""
        periodo = request.args.get("periodo", 2, type=int)
        alunos = aluno_service.listar_por_periodo(periodo)
        todos = alunos
        alertas = []

        for aluno in alunos:
            freq  = indicador_service.calcular_frequencia_media(aluno, periodo)
            media = indicador_service.calcular_media_geral(aluno, periodo)
            cv    = indicador_service.calcular_coeficiente_variacao(aluno, periodo)
            zscore = indicador_service.calcular_zscore_aluno(todos, aluno.id, periodo)

            # Sem dados no período — ignora
            if freq == 0.0 and media == 0.0:
                continue

            # Frequência crítica (abaixo do mínimo legal)
            if freq < 75:
                alertas.append({
                    "nivel": "critico",
                    "aluno_id": aluno.id,
                    "aluno_nome": aluno.nome,
                    "tipo": "frequencia",
                    "mensagem": f"Frequência crítica: {freq:.1f}% — abaixo do mínimo de 75%",
                    "acao": "Contato urgente e elaboração de plano de reposição",
                })
            elif freq < 82:
                alertas.append({
                    "nivel": "alto",
                    "aluno_id": aluno.id,
                    "aluno_nome": aluno.nome,
                    "tipo": "frequencia",
                    "mensagem": f"Frequência em atenção: {freq:.1f}% — próxima do limite mínimo",
                    "acao": "Monitorar assiduidade e conversar com o aluno",
                })

            # Desempenho abaixo do mínimo de aprovação
            if media < 5.0:
                alertas.append({
                    "nivel": "critico",
                    "aluno_id": aluno.id,
                    "aluno_nome": aluno.nome,
                    "tipo": "desempenho",
                    "mensagem": f"Média abaixo do mínimo: {media:.2f} — risco de reprovação",
                    "acao": "Agendar tutoria individual e reforço nas disciplinas críticas",
                })
            elif media < 6.5:
                alertas.append({
                    "nivel": "alto",
                    "aluno_id": aluno.id,
                    "aluno_nome": aluno.nome,
                    "tipo": "desempenho",
                    "mensagem": f"Média em zona de risco: {media:.2f}",
                    "acao": "Incentivar participação e verificar dificuldades específicas",
                })

            # Desempenho inconsistente entre disciplinas
            if cv is not None and cv > 35:
                alertas.append({
                    "nivel": "medio",
                    "aluno_id": aluno.id,
                    "aluno_nome": aluno.nome,
                    "tipo": "inconsistencia",
                    "mensagem": f"Desempenho muito irregular entre disciplinas (CV: {cv:.1f}%)",
                    "acao": "Identificar disciplinas com maior dificuldade e direcionar suporte",
                })

            # Desempenho muito abaixo da média da turma
            if zscore is not None and zscore < -1.5:
                alertas.append({
                    "nivel": "alto",
                    "aluno_id": aluno.id,
                    "aluno_nome": aluno.nome,
                    "tipo": "comparativo",
                    "mensagem": f"Desempenho {abs(zscore):.1f} desvios abaixo da média da turma (z={zscore:.2f})",
                    "acao": "Avaliar necessidade de acompanhamento pedagógico especializado",
                })

        ordem = {"critico": 0, "alto": 1, "medio": 2}
        alertas.sort(key=lambda a: ordem.get(a["nivel"], 3))

        return jsonify({"periodo": periodo, "total": len(alertas), "alertas": alertas}), 200


    # ------------------------------------------------------------------ #
    # Módulo de Importação e Painel Acadêmico (Database-driven)
    # ------------------------------------------------------------------ #

    @app.route("/api/importar/analisar", methods=["POST"])
    @requer_autenticacao
    @requer_permissao("escrita")
    def analisar_planilha():
        dados = request.get_json() or {}
        url = dados.get("url", "").strip()
        if not url:
            return jsonify({"erro": "A URL da planilha é obrigatória."}), 400

        # Extrai o ID da planilha do Google Sheets
        match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
        if match:
            spreadsheet_id = match.group(1)
            csv_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv"
        else:
            csv_url = url

        try:
            import requests
            import pandas as pd

            response = requests.get(csv_url, timeout=30)
            if response.status_code != 200:
                return jsonify({"erro": f"Não foi possível acessar a planilha. Status: {response.status_code}. Certifique-se de que a planilha é pública."}), 400

            # Salva o arquivo temporário
            import_id = str(uuid.uuid4())
            temp_dir = os.path.join(app.root_path, "..", "data")
            os.makedirs(temp_dir, exist_ok=True)
            temp_path = os.path.join(temp_dir, f"temp_import_{import_id}.csv")

            with open(temp_path, "wb") as f:
                f.write(response.content)

            # Lê o CSV para extrair colunas e prévia
            df = pd.read_csv(temp_path, sep=None, engine="python", nrows=5, dtype=str)
            columns = [c.strip() for c in df.columns]

            # Gera prévia
            preview_rows = df.fillna("").to_dict(orient="records")

            # Mapeamento automático inteligente
            system_columns = {
                "ID_ALUNO": ["pessoa", "id_aluno", "matricula", "cod_aluno", "aluno_id"],
                "NOME_ALUNO": ["nome_aluno", "nome_completo", "nome_estudante"],
                "COD_CURSO": ["cod_curso", "codigo_curso", "curso_cod"],
                "NOME_CURSO": ["nome_curso", "curso", "nome"],
                "ANO": ["ano", "ano_letivo", "periodo"],
                "SEMESTRE": ["semestre", "sem", "periodo_letivo"],
                "SERIE": ["serie", "periodo", "semestre_curso"],
                "ANO_INGRESSO": ["ano_ingresso", "periodo_ingresso", "ingresso_ano"],
                "SEM_INGRESSO": ["sem_ingresso", "ingresso_sem"],
                "COD_DISCIPLINA": ["cod_disciplina", "codigo_disciplina", "disciplina"],
                "NOME_DISCIPLINA": ["nome_disciplina", "disciplina_nome", "nome_disciplina"],
                "TURMA": ["turma", "cod_turma", "codigo_turma"],
                "SITUACAO": ["situacao", "situacao_hist", "status", "resultado"],
                "RISCO_ACADEMICO": ["risco", "risco_academico", "nivel_risco"],
                "VA1": ["va1", "nota_va1", "nota1"],
                "VA2": ["va2", "nota_va2", "nota2"],
                "VA3": ["va3", "nota_va3", "nota3"],
                "MEDIA_FINAL": ["media_final", "media_nota", "media", "nota_final"],
                "MEDIA_CALCULADA": ["media_calculada", "media_calc"],
                "UNIDADE": ["unidade", "unidade_fisica", "campus"],
                "CIDADE": ["cidade", "municipio"],
                "ESTADO": ["estado", "uf"],
                "SEXO": ["sexo", "genero"],
                "DATA_NASCIMENTO": ["data_nascimento", "nascimento", "nasc"],
                "IDADE": ["idade"],
                "PROUNI": ["prouni", "bolsa_prouni"],
                "FIES": ["fies", "financiamento_fies"],
                "BOLSA": ["bolsa", "outras_bolsas", "desconto_funcionario"],
                "PERC_PRESENCA": ["perc_presenca", "presenca", "frequencia", "freq"]
            }

            sugestoes = {}
            for sys_col, aliases in system_columns.items():
                best_col = None
                best_score = -1
                for col in columns:
                    col_lower = col.lower().strip()
                    if col_lower == sys_col.lower() or col_lower in aliases:
                        score = 10
                    elif any(col_lower == alias.replace("_", " ") for alias in aliases):
                        score = 8
                    elif any(alias == col_lower for alias in aliases):
                        score = 7
                    elif any(alias in col_lower for alias in aliases):
                        score = 5
                    else:
                        score = 0
                    
                    if score > best_score and score > 0:
                        best_score = score
                        best_col = col
                if best_col:
                    sugestoes[sys_col] = best_col
                else:
                    if sys_col == "NOME_CURSO" and "NOME" in columns:
                        sugestoes["NOME_CURSO"] = "NOME"
                    elif sys_col == "ID_ALUNO" and "PESSOA" in columns:
                        sugestoes["ID_ALUNO"] = "PESSOA"

            return jsonify({
                "import_id": import_id,
                "columns": columns,
                "sugestoes": sugestoes,
                "preview": preview_rows
            }), 200

        except Exception as e:
            return jsonify({"erro": f"Erro ao processar planilha: {str(e)}"}), 500


    @app.route("/api/importar/confirmar", methods=["POST"])
    @requer_autenticacao
    @requer_permissao("escrita")
    def confirmar_importacao():
        dados = request.get_json() or {}
        import_id = dados.get("import_id")
        mappings = dados.get("mappings")

        if not import_id or not mappings:
            return jsonify({"erro": "Parâmetros import_id e mappings são obrigatórios."}), 400

        temp_path = os.path.join(app.root_path, "..", "data", f"temp_import_{import_id}.csv")
        if not os.path.exists(temp_path):
            return jsonify({"erro": "Arquivo de importação não encontrado ou já processado."}), 404

        try:
            import pandas as pd
            import sqlite3
            import json

            df = pd.read_csv(temp_path, sep=None, engine="python", dtype=str)
            df.columns = [c.strip() for c in df.columns]

            required_sys_cols = ["ID_ALUNO", "NOME_CURSO", "TURMA", "ANO", "SEMESTRE", "NOME_DISCIPLINA", "MEDIA_FINAL"]
            for rc in required_sys_cols:
                if rc not in mappings or not mappings[rc] or mappings[rc] not in df.columns:
                    return jsonify({"erro": f"Mapeamento da coluna obrigatória '{rc}' ausente ou incorreto."}), 400

            def _to_float(val):
                if pd.isna(val) or str(val).strip() == "":
                    return None
                try:
                    v = float(str(val).replace(",", ".").strip())
                    if v > 10.0:
                        v = v / 10.0
                    return round(v, 2)
                except Exception:
                    return None

            def _to_int(val):
                if pd.isna(val) or str(val).strip() == "":
                    return None
                try:
                    return int(float(str(val).strip()))
                except Exception:
                    return None

            def _to_str(val):
                if pd.isna(val):
                    return None
                s = str(val).strip()
                return s if s != "" else None

            def _parse_frequency(val, situacao):
                if pd.isna(val) or str(val).strip() == "":
                    if situacao == "Rep_Freq":
                        return 70.0
                    else:
                        return 95.0
                try:
                    v = float(str(val).replace(",", ".").strip())
                    if v <= 1.0:
                        return round(v * 100.0, 1)
                    elif v <= 10.0:
                        return round(v * 10.0, 1)
                    else:
                        return round(v, 1)
                except Exception:
                    return 95.0

            conn = get_connection()
            conn.execute("PRAGMA foreign_keys = OFF")

            try:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM registros_academicos")
                cursor.execute("DELETE FROM disciplinas")
                cursor.execute("DELETE FROM alunos")
                cursor.execute("DELETE FROM indicadores")

                id_aluno_col = mappings["ID_ALUNO"]
                nome_aluno_col = mappings.get("NOME_ALUNO")
                nome_curso_col = mappings["NOME_CURSO"]
                cod_curso_col = mappings.get("COD_CURSO")
                ano_ingresso_col = mappings.get("ANO_INGRESSO")
                sem_ingresso_col = mappings.get("SEM_INGRESSO")
                unidade_col = mappings.get("UNIDADE")
                cidade_col = mappings.get("CIDADE")
                estado_col = mappings.get("ESTADO")
                sexo_col = mappings.get("SEXO")
                data_nasc_col = mappings.get("DATA_NASCIMENTO")
                idade_col = mappings.get("IDADE")
                prouni_col = mappings.get("PROUNI")
                fies_col = mappings.get("FIES")
                bolsa_col = mappings.get("BOLSA")

                unique_students_df = df.drop_duplicates(subset=[id_aluno_col])
                alunos_registros = []
                student_uuid_map = {}

                for _, row in unique_students_df.iterrows():
                    matricula = str(row[id_aluno_col])
                    student_id = str(uuid.uuid4())
                    student_uuid_map[matricula] = student_id

                    if nome_aluno_col and nome_aluno_col in row and not pd.isna(row[nome_aluno_col]):
                        nome_aluno = _to_str(row[nome_aluno_col])
                    else:
                        nome_aluno = f"Aluno {matricula}"

                    alunos_registros.append({
                        "id": student_id,
                        "matricula": matricula,
                        "nome": nome_aluno,
                        "curso": _to_str(row.get(nome_curso_col)) if nome_curso_col in row else None,
                        "cod_curso": _to_int(row.get(cod_curso_col)) if cod_curso_col and cod_curso_col in row else None,
                        "periodo": None,
                        "ano_ingresso": _to_int(row.get(ano_ingresso_col)) if ano_ingresso_col and ano_ingresso_col in row else None,
                        "sem_ingresso": _to_int(row.get(sem_ingresso_col)) if sem_ingresso_col and sem_ingresso_col in row else None,
                        "unidade": _to_str(row.get(unidade_col)) if unidade_col and unidade_col in row else None,
                        "cidade": _to_str(row.get(cidade_col)) if cidade_col and cidade_col in row else None,
                        "estado": _to_str(row.get(estado_col)) if estado_col and estado_col in row else None,
                        "sexo": _to_str(row.get(sexo_col)) if sexo_col and sexo_col in row else None,
                        "data_nascimento": _to_str(row.get(data_nasc_col)) if data_nasc_col and data_nasc_col in row else None,
                        "idade": _to_int(row.get(idade_col)) if idade_col and idade_col in row else None,
                        "prouni": _to_str(row.get(prouni_col)) if prouni_col and prouni_col in row else "N",
                        "fies": _to_str(row.get(fies_col)) if fies_col and fies_col in row else "N",
                        "bolsa": _to_str(row.get(bolsa_col)) if bolsa_col and bolsa_col in row else "N",
                        "email": None,
                        "ativo": 1,
                        "notas": "[]",
                        "frequencias": "[]"
                    })

                for a in alunos_registros:
                    cursor.execute("""
                        INSERT INTO alunos 
                        (id, matricula, nome, curso, cod_curso, periodo, ano_ingresso, sem_ingresso, 
                         unidade, cidade, estado, sexo, data_nascimento, idade, prouni, fies, bolsa, email, ativo, notas, frequencias)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (a["id"], a["matricula"], a["nome"], a["curso"], a["cod_curso"], a["periodo"], a["ano_ingresso"], a["sem_ingresso"],
                          a["unidade"], a["cidade"], a["estado"], a["sexo"], a["data_nascimento"], a["idade"], a["prouni"], a["fies"], a["bolsa"],
                          a["email"], a["ativo"], a["notas"], a["frequencias"]))

                cod_disciplina_col = mappings.get("COD_DISCIPLINA")
                nome_disciplina_col = mappings["NOME_DISCIPLINA"]

                if cod_disciplina_col:
                    disc_df = df.drop_duplicates(subset=[cod_disciplina_col])
                    for _, row in disc_df.iterrows():
                        cod_d = _to_int(row[cod_disciplina_col])
                        if cod_d is None:
                            continue
                        nome_d = _to_str(row.get(nome_disciplina_col)) or f"Disciplina {cod_d}"
                        cursor.execute("""
                            INSERT OR IGNORE INTO disciplinas (cod_disciplina, nome, carga_horaria)
                            VALUES (?,?,?)
                        """, (cod_d, nome_d, None))
                else:
                    unique_discs = df[nome_disciplina_col].unique()
                    for i, nome_d in enumerate(unique_discs, start=1000):
                        if pd.isna(nome_d) or str(nome_d).strip() == "":
                            continue
                        cursor.execute("""
                            INSERT OR IGNORE INTO disciplinas (cod_disciplina, nome, carga_horaria)
                            VALUES (?,?,?)
                        """, (i, str(nome_d).strip(), None))

                disc_rows = cursor.execute("SELECT cod_disciplina, nome FROM disciplinas").fetchall()
                disc_name_to_cod = {r["nome"]: r["cod_disciplina"] for r in disc_rows}

                turma_col = mappings["TURMA"]
                ano_col = mappings["ANO"]
                semestre_col = mappings["SEMESTRE"]
                serie_col = mappings.get("SERIE")
                situacao_col = mappings.get("SITUACAO")
                risco_col = mappings.get("RISCO_ACADEMICO")
                va1_col = mappings.get("VA1")
                va2_col = mappings.get("VA2")
                va3_col = mappings.get("VA3")
                media_final_col = mappings["MEDIA_FINAL"]
                media_calc_col = mappings.get("MEDIA_CALCULADA")
                freq_col = mappings.get("PERC_PRESENCA")

                student_notas = {a["id"]: [] for a in alunos_registros}
                student_freqs = {a["id"]: [] for a in alunos_registros}
                student_series = {a["id"]: set() for a in alunos_registros}

                registros_importados = 0
                for _, row in df.iterrows():
                    matricula = str(row[id_aluno_col])
                    aluno_id = student_uuid_map.get(matricula)
                    if not aluno_id:
                        continue

                    nome_d = _to_str(row.get(nome_disciplina_col))
                    if not nome_d:
                        continue

                    if cod_disciplina_col:
                        cod_d = _to_int(row[cod_disciplina_col])
                    else:
                        cod_d = disc_name_to_cod.get(nome_d)

                    turma = _to_str(row.get(turma_col))
                    ano = _to_int(row.get(ano_col)) or 2024
                    semestre = _to_int(row.get(semestre_col)) or 1

                    serie = _to_int(row.get(serie_col)) if serie_col else 1
                    if serie:
                        student_series[aluno_id].add(serie)
                    else:
                        serie = 1

                    sit_raw = _to_str(row.get(situacao_col)) if situacao_col else "Aprovado"
                    if sit_raw in ("Aprovado", "Rep_Nota", "Rep_Freq"):
                        situacao = sit_raw
                    elif "rep" in sit_raw.lower() and "freq" in sit_raw.lower():
                        situacao = "Rep_Freq"
                    elif "rep" in sit_raw.lower():
                        situacao = "Rep_Nota"
                    elif "disp" in sit_raw.lower():
                        situacao = "Aprovado"
                    else:
                        situacao = "Aprovado"

                    risco_raw = _to_str(row.get(risco_col)) if risco_col else None
                    if risco_raw in ("Baixo", "Medio", "Alto"):
                        risco = risco_raw
                    else:
                        risco = None

                    va1 = _to_float(row.get(va1_col)) if va1_col else None
                    va2 = _to_float(row.get(va2_col)) if va2_col else None
                    va3 = _to_float(row.get(va3_col)) if va3_col else None
                    media_final = _to_float(row.get(media_final_col))
                    media_calc = _to_float(row.get(media_calc_col)) if media_calc_col else media_final

                    freq_val = _parse_frequency(row.get(freq_col) if freq_col else None, situacao)

                    if not risco:
                        if situacao == "Aprovado" and media_final is not None and media_final >= 6.0:
                            risco = "Baixo"
                        elif (media_final is not None and media_final < 4.0) or freq_val < 70.0:
                            risco = "Alto"
                        else:
                            risco = "Medio"

                    cursor.execute("""
                        INSERT OR IGNORE INTO registros_academicos 
                        (aluno_id, cod_disciplina, nome_disciplina, turma, serie, ano, semestre,
                         va1, va2, va3, media_final, media_calculada, situacao, risco_academico)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (aluno_id, cod_d, nome_d, turma, serie, ano, semestre,
                          va1, va2, va3, media_final, media_calc, situacao, risco))

                    registros_importados += 1

                    if media_final is not None:
                        student_notas[aluno_id].append({
                            "disciplina": nome_d,
                            "nota": media_final,
                            "periodo": serie
                        })
                    if freq_val is not None:
                        student_freqs[aluno_id].append({
                            "disciplina": nome_d,
                            "percentual": freq_val,
                            "periodo": serie
                        })

                for a in alunos_registros:
                    a_id = a["id"]
                    series_set = student_series[a_id]
                    max_periodo = max(series_set) if series_set else 1

                    notas_json = json.dumps(student_notas[a_id], ensure_ascii=False)
                    freqs_json = json.dumps(student_freqs[a_id], ensure_ascii=False)

                    cursor.execute("""
                        UPDATE alunos 
                        SET periodo = ?, notas = ?, frequencias = ?
                        WHERE id = ?
                    """, (max_periodo, notas_json, freqs_json, a_id))

                conn.commit()

                # 4. Gera indicadores para todos os alunos e períodos
                cursor = conn.cursor()
                all_alunos_db = cursor.execute("SELECT * FROM alunos").fetchall()
                todos_indicadores = []
                for a_row in all_alunos_db:
                    a_dict = dict(a_row)
                    a_dict["notas"] = json.loads(a_dict["notas"]) if a_dict["notas"] else []
                    a_dict["frequencias"] = json.loads(a_dict["frequencias"]) if a_dict["frequencias"] else []
                    aluno_obj = Aluno.from_dict(a_dict)
                    a_periodos = set(n["periodo"] for n in aluno_obj.get_notas())
                    if not a_periodos:
                        a_periodos = {1}
                    for p in a_periodos:
                        inds = indicador_service.gerar_indicadores_aluno(aluno_obj, p, salvar=False)
                        for ind in inds:
                            todos_indicadores.append((ind.id, ind.to_dict()))

                # Salva todos os indicadores de uma vez só!
                repo_indicadores.salvar_lote(todos_indicadores)

            except Exception as db_err:
                try:
                    conn.rollback()
                except Exception:
                    pass
                raise db_err
            finally:
                conn.execute("PRAGMA foreign_keys = ON")
                conn.close()

            if os.path.exists(temp_path):
                os.remove(temp_path)

            token = request.headers.get("Authorization", "").replace("Bearer ", "")
            registrar_auditoria("importar_planilha_google", token, {"alunos_importados": len(alunos_registros), "registros_importados": registros_importados})

            return jsonify({
                "sucesso": True,
                "mensagem": "Importação concluída com sucesso!",
                "alunos_importados": len(alunos_registros),
                "registros_importados": registros_importados
            }), 200

        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({"erro": f"Erro durante a importação: {str(e)}"}), 500


    @app.route("/turmas", methods=["GET"])
    @requer_autenticacao
    def listar_turmas():
        try:
            import json
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute("SELECT DISTINCT turma FROM registros_academicos WHERE turma IS NOT NULL AND turma != ''")
            turmas_raw = [r[0] for r in cursor.fetchall()]

            turmas = []
            for t_code in turmas_raw:
                cursor.execute("""
                    SELECT ra.aluno_id, ra.media_final, ra.situacao, ra.risco_academico, a.curso, ra.serie
                    FROM registros_academicos ra
                    JOIN alunos a ON ra.aluno_id = a.id
                    WHERE ra.turma = ?
                """, (t_code,))
                rows = cursor.fetchall()
                if not rows:
                    continue

                student_records = {}
                for r in rows:
                    aluno_id = r[0]
                    media = r[1]
                    situacao = r[2]
                    risco = r[3]
                    curso = r[4]
                    serie = r[5]

                    if aluno_id not in student_records:
                        student_records[aluno_id] = {
                            "medias": [],
                            "situacoes": [],
                            "riscos": [],
                            "cursos": [],
                            "series": []
                        }
                    if media is not None:
                        student_records[aluno_id]["medias"].append(media)
                    if situacao:
                        student_records[aluno_id]["situacoes"].append(situacao)
                    if risco:
                        student_records[aluno_id]["riscos"].append(risco)
                    if curso:
                        student_records[aluno_id]["cursos"].append(curso)
                    if serie is not None:
                        student_records[aluno_id]["series"].append(serie)

                total_students = len(student_records)
                all_medias = [m for s in student_records.values() for m in s["medias"]]
                avg_media = sum(all_medias) / len(all_medias) if all_medias else 0.0

                # Frequências
                all_freqs = []
                cursor.execute("""
                    SELECT a.frequencias, ra.serie
                    FROM registros_academicos ra
                    JOIN alunos a ON ra.aluno_id = a.id
                    WHERE ra.turma = ?
                """, (t_code,))
                freq_rows = cursor.fetchall()
                for f_row in freq_rows:
                    freqs_json = f_row[0]
                    periodo = f_row[1]
                    try:
                        freqs_list = json.loads(freqs_json) if freqs_json else []
                        for f in freqs_list:
                            if f.get("periodo") == periodo and "percentual" in f:
                                all_freqs.append(f["percentual"])
                    except Exception:
                        pass

                avg_freq = sum(all_freqs) / len(all_freqs) if all_freqs else 95.0

                risk_counts = {"baixo": 0, "medio": 0, "alto": 0, "critico": 0}
                for s_id, s_data in student_records.items():
                    s_riscos = s_data["riscos"]
                    if s_riscos:
                        most_common = max(set(s_riscos), key=s_riscos.count).lower()
                        if most_common == "critico":
                            risk_counts["critico"] += 1
                        elif most_common == "alto":
                            risk_counts["alto"] += 1
                        elif most_common == "medio":
                            risk_counts["medio"] += 1
                        else:
                            risk_counts["baixo"] += 1
                    else:
                        risk_counts["baixo"] += 1

                alerts_count = risk_counts["alto"] + risk_counts["critico"]

                all_situacoes = [s for sd in student_records.values() for s in sd["situacoes"]]
                aprovados = sum(1 for s in all_situacoes if s == "Aprovado")
                taxa_aprov = (aprovados / len(all_situacoes) * 100) if all_situacoes else 100.0

                all_courses = [c for sd in student_records.values() for c in sd["cursos"]]
                most_common_course = max(set(all_courses), key=all_courses.count) if all_courses else "Engenharia de Software"

                all_series = [s for sd in student_records.values() for s in sd["series"]]
                most_common_period = max(set(all_series), key=all_series.count) if all_series else 1

                turno = "Noturno"
                if "M" in t_code or "MAT" in t_code.upper():
                    turno = "Matutino"
                elif "V" in t_code or "VESP" in t_code.upper():
                    turno = "Vespertino"

                turmas.append({
                    "id": t_code,
                    "nome": f"Turma {t_code}",
                    "curso": most_common_course,
                    "periodo": most_common_period,
                    "turno": turno,
                    "total_alunos": total_students,
                    "media_turma": round(avg_media, 2),
                    "freq_media": round(avg_freq, 1),
                    "alunos_em_alerta": alerts_count,
                    "taxa_aprovacao": round(taxa_aprov, 1),
                    "distribuicao_risco": risk_counts
                })

            conn.close()
            return jsonify(turmas), 200

        except Exception as e:
            return jsonify({"erro": str(e)}), 500


    @app.route("/turmas/<turma_id>/relatorio", methods=["GET"])
    @requer_autenticacao
    def relatorio_detalhado_turma(turma_id):
        try:
            import json
            import statistics
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT ra.aluno_id, ra.serie, a.curso
                FROM registros_academicos ra
                JOIN alunos a ON ra.aluno_id = a.id
                WHERE ra.turma = ?
            """, (turma_id,))
            rows = cursor.fetchall()
            if not rows:
                conn.close()
                return jsonify({"erro": "Turma não encontrada."}), 404

            aluno_ids = list(set(r[0] for r in rows))
            series = [r[1] for r in rows if r[1] is not None]
            periodo = max(set(series), key=series.count) if series else 1
            cursos = [r[2] for r in rows if r[2]]
            curso = max(set(cursos), key=cursos.count) if cursos else "Engenharia de Software"

            turno = "Noturno"
            if "M" in turma_id or "MAT" in turma_id.upper():
                turno = "Matutino"
            elif "V" in turma_id or "VESP" in turma_id.upper():
                turno = "Vespertino"

            turma_info = {
                "id": turma_id,
                "nome": f"Turma {turma_id}",
                "curso": curso,
                "periodo": periodo,
                "turno": turno
            }

            db_alunos = []
            for a_id in aluno_ids:
                a_data = repo_alunos.buscar(a_id)
                if a_data:
                    db_alunos.append(Aluno.from_dict(a_data))

            alunos_metrics = []
            for aluno in db_alunos:
                media = indicador_service.calcular_media_geral(aluno, periodo)
                freq = indicador_service.calcular_frequencia_media(aluno, periodo)
                if freq == 0.0:
                    cursor.execute("""
                        SELECT situacao FROM registros_academicos 
                        WHERE aluno_id = ? AND turma = ? LIMIT 1
                    """, (aluno.id, turma_id))
                    sit_row = cursor.fetchone()
                    if sit_row and sit_row[0] == "Rep_Freq":
                        freq = 70.0
                    else:
                        freq = 95.0

                iaa = indicador_service.calcular_iaa(aluno, periodo)
                irp = indicador_service.calcular_irp(aluno, periodo)
                risco = indicador_service.classificar_risco(aluno, periodo).nivel.value

                zscore = indicador_service.calcular_zscore_aluno(db_alunos, aluno.id, periodo)
                percentil = indicador_service.calcular_percentil_aluno(db_alunos, aluno.id, periodo)

                alunos_metrics.append({
                    "aluno_id": aluno.id,
                    "nome": aluno.nome,
                    "matricula": aluno.matricula,
                    "media_geral": media,
                    "frequencia_media": freq,
                    "iaa": iaa,
                    "irp": irp,
                    "risco": risco,
                    "zscore": zscore,
                    "percentil": percentil
                })

            total = len(alunos_metrics)
            medias = [a["media_geral"] for a in alunos_metrics]
            freqs = [a["frequencia_media"] for a in alunos_metrics]
            avg_media = sum(medias) / len(medias) if medias else 0.0
            avg_freq = sum(freqs) / len(freqs) if freqs else 0.0
            em_alerta = sum(1 for a in alunos_metrics if a["risco"] in ("alto", "critico"))
            criticos = sum(1 for a in alunos_metrics if a["risco"] == "critico")

            cursor.execute("""
                SELECT COUNT(*) FROM registros_academicos
                WHERE turma = ? AND situacao = 'Aprovado'
            """, (turma_id,))
            aprovados = cursor.fetchone()[0]
            cursor.execute("""
                SELECT COUNT(*) FROM registros_academicos
                WHERE turma = ?
            """, (turma_id,))
            total_records = cursor.fetchone()[0] or 1
            taxa_aprov = (aprovados / total_records) * 100

            stats = {
                "media": round(avg_media, 2),
                "freq": round(avg_freq, 1),
                "em_alerta": em_alerta,
                "criticos": criticos,
                "taxa_aprovacao": round(taxa_aprov, 1)
            }

            pontos = [{"nome": a["nome"], "frequencia": a["frequencia_media"], "media": a["media_geral"]} for a in alunos_metrics]

            cursor.execute("""
                SELECT DISTINCT cod_disciplina, nome_disciplina FROM registros_academicos WHERE turma = ?
            """, (turma_id,))
            disc_rows = cursor.fetchall()
            disciplinas = []
            for d_row in disc_rows:
                cod_d = d_row[0]
                nome_d = d_row[1]

                cursor.execute("""
                    SELECT media_final, situacao FROM registros_academicos
                    WHERE turma = ? AND nome_disciplina = ?
                """, (turma_id, nome_d))
                d_records = cursor.fetchall()
                if not d_records:
                    continue
                d_medias = [r[0] for r in d_records if r[0] is not None]
                d_avg = sum(d_medias) / len(d_medias) if d_medias else 0.0

                d_medias_sorted = sorted(d_medias)
                n_m = len(d_medias_sorted)
                if n_m == 0:
                    d_med = 0.0
                elif n_m % 2 == 1:
                    d_med = d_medias_sorted[n_m // 2]
                else:
                    d_med = (d_medias_sorted[n_m // 2 - 1] + d_medias_sorted[n_m // 2]) / 2.0

                d_stdev = 0.0
                if len(d_medias) >= 2:
                    d_stdev = statistics.stdev(d_medias)

                d_min = min(d_medias) if d_medias else 0.0
                d_max = max(d_medias) if d_medias else 0.0
                d_total = len(d_records)
                d_aprovados = sum(1 for r in d_records if r[1] == 'Aprovado')
                d_taxa_aprov = (d_aprovados / d_total * 100) if d_total else 100.0

                disciplinas.append({
                    "disciplina": nome_d,
                    "media": round(d_avg, 2),
                    "mediana": round(d_med, 2),
                    "desvio_padrao": round(d_stdev, 2),
                    "nota_min": round(d_min, 2),
                    "nota_max": round(d_max, 2),
                    "total_alunos": d_total,
                    "taxa_aprovacao": round(d_taxa_aprov, 1)
                })

            conn.close()
            return jsonify({
                "turma": turma_info,
                "total": total,
                "alunos": alunos_metrics,
                "stats": stats,
                "pontos_dispersao": pontos,
                "disciplinas": disciplinas
            }), 200

        except Exception as e:
            return jsonify({"erro": str(e)}), 500

    @app.route("/api/charts/turma/<turma_id>", methods=["GET"])
    @requer_autenticacao
    def get_turma_charts(turma_id):
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT DISTINCT a.curso, ra.serie FROM registros_academicos ra JOIN alunos a ON ra.aluno_id = a.id WHERE ra.turma = ?", (turma_id,))
            turma_row = cursor.fetchone()
            turma_nome = f"Turma {turma_id}" + (f" ({turma_row[0]} - {turma_row[1]}º Período)" if turma_row else "")

            cursor.execute("SELECT a.nome, ra.media_final, a.frequencias, ra.serie FROM registros_academicos ra JOIN alunos a ON ra.aluno_id = a.id WHERE ra.turma = ?", (turma_id,))
            pontos_dispersao = []
            for r in cursor.fetchall():
                media, freqs_json, serie = r[1], r[2], r[3]
                freq_val = 95.0
                if freqs_json:
                    try:
                        freq_item = next((f for f in json.loads(freqs_json) if f.get("periodo") == serie), None)
                        if freq_item: freq_val = freq_item.get("percentual", 95.0)
                    except: pass
                if media is not None: pontos_dispersao.append({"nome": r[0], "media": media, "frequencia": freq_val})

            medias_turma = [p['media'] for p in pontos_dispersao]
            
            cursor.execute("SELECT ra.nome_disciplina, ra.media_final FROM registros_academicos ra WHERE ra.turma = ?", (turma_id,))
            disciplinas_raw = {}
            for r in cursor.fetchall():
                if r[0] not in disciplinas_raw: disciplinas_raw[r[0]] = []
                if r[1] is not None: disciplinas_raw[r[0]].append(r[1])
            disciplinas_data = [{"disciplina": nome, "medias": medias} for nome, medias in disciplinas_raw.items()]

            script_scatter, div_scatter = BokehService.gerar_dispersao_nota_frequencia(pontos_dispersao, turma_nome)
            script_hist, div_hist = BokehService.gerar_histograma_desempenho(medias_turma, turma_nome)
            script_box, div_box = BokehService.gerar_boxplot_disciplinas(disciplinas_data, turma_nome)

            return jsonify({
                "scatter": {"script": script_scatter, "div": div_scatter},
                "histogram": {"script": script_hist, "div": div_hist},
                "boxplot": {"script": script_box, "div": div_box},
            })
        except Exception as e:
            return jsonify({"erro": f"Erro ao gerar gráficos: {str(e)}"}), 500

    # ------------------------------------------------------------------ #
    # Healthcheck
    # ------------------------------------------------------------------ #

    @app.route("/", methods=["GET"])
    def healthcheck():
        return redirect(url_for("login_page"))


    # ------------------------------------------------------------------ #
    # Tratamento global de erros
    # ------------------------------------------------------------------ #

    @app.errorhandler(404)
    def nao_encontrado(e):
        return jsonify({"erro": "Rota não encontrada."}), 404

    @app.errorhandler(405)
    def metodo_nao_permitido(e):
        return jsonify({"erro": "Método HTTP não permitido para esta rota."}), 405

    @app.errorhandler(500)
    def erro_interno(e):
        return jsonify({"erro": "Erro interno do servidor."}), 500


    if __name__ == "__main__":
        port = int(os.environ.get("PORT", 5000))
        app.run(host="0.0.0.0", debug=True, port=port)


    @app.route("/api/charts/geral", methods=["GET"])
    @requer_autenticacao
    def get_charts_geral():
        periodo = request.args.get("periodo", 1, type=int)
        conn = get_connection()
        cur = conn.cursor()

        # Apply RBAC filtering scope
        query_scope = ""
        params = []
        if request.perfil_logado in ["professor", "visualizador"]:
            query_scope = " AND ra.cod_turma IN (SELECT cod_turma FROM turmas LIMIT 2) " # Fake scope constraint for mock

        # Notas Distribution
        cur.execute(f"SELECT nota_final FROM registros_academicos ra WHERE ano = 2023 AND semestre = ? {query_scope}", [periodo] + params)
        notas = [r[0] for r in cur.fetchall() if r[0] is not None]

        s_hist, d_hist = BokehService.gerar_histograma_desempenho(notas)

        # Scatter Nota x Freq
        cur.execute(f"SELECT nota_final, freq_presenca FROM registros_academicos ra WHERE ano = 2023 AND semestre = ? {query_scope}", [periodo] + params)
        dados_scatter = [{"nota": r[0], "freq": r[1]} for r in cur.fetchall() if r[0] is not None and r[1] is not None]

        s_scat, d_scat = BokehService.gerar_dispersao_nota_frequencia(dados_scatter)

        # Boxplot por turma
        cur.execute(f"SELECT t.nome_disciplina, ra.nota_final FROM registros_academicos ra JOIN turmas t ON ra.cod_turma = t.cod_turma WHERE ra.ano = 2023 AND ra.semestre = ? {query_scope}", [periodo] + params)
        box_data = {}
        for r in cur.fetchall():
            turma_nome = r[0]
            nota = r[1]
            if nota is not None:
                if turma_nome not in box_data:
                    box_data[turma_nome] = []
                box_data[turma_nome].append(nota)

        s_box, d_box = BokehService.gerar_boxplot_variabilidade(box_data)

        conn.close()

        return jsonify({
            "hist": {"script": s_hist, "div": d_hist},
            "scatter": {"script": s_scat, "div": d_scat},
            "boxplot": {"script": s_box, "div": d_box}
        })

    return app

if __name__ == "__main__":
    app = create_app()
    app.run(host="0.0.0.0", debug=True, port=int(os.environ.get("PORT", 5000)))
